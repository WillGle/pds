import os
import time
import re
import shutil
import subprocess
import datetime
import hashlib
import base64
import io
import gc
import glob
from queue import Queue
import pandas as pd
import streamlit as st
from concurrent.futures import ThreadPoolExecutor, as_completed



from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException


def _k(cipher_b64):
    k_str = os.environ.get('PDS_KEY', '')
    try:
        if hasattr(st, "secrets") and "PDS_KEY" in st.secrets:
            k_str = st.secrets["PDS_KEY"]
    except Exception:
        pass
    kb = hashlib.sha256(k_str.encode('utf-8')).digest()
    raw = base64.b64decode(cipher_b64)
    return bytes([x ^ kb[i % len(kb)] for i, x in enumerate(raw)]).decode('utf-8')


S_CHROMIUM = _k("ISSfGrEeRVQ=")
S_CHROMIUM_BROWSER = _k("ISSfGrEeRVRIxtT4LoFEWA==")
S_GOOGLE_CHROME = _k("JSOCErASHVoN1sn6PA==")
S_USR_CHROMIUM = _k("bTmeB/MVWVdKx87lNp9IXxk=")
S_CHROMEDRIVER = _k("ISSfGrESVEsM0sPl")
S_CHROMIUM_DRIVER = _k("ISSfGrEeRVRIwNT+L5dT")
S_USR_CHROMEDRIVER = _k("bTmeB/MVWVdKx87lNp9ETgbvp5mw")
S_REGEX_ID = _k("anPXA+ELH0sAwcrkZt1dBQLvtZmtqkCbmuzRNGSAvBMmZ8Q=")
S_POPUP_JS = _k("NC2fVb8bX0oA5tL5ec8BThvlpJGntxvJnbbIMCD68SMnL5karl8XXQzS/fYrm0AHGOezma7kTSN8AB4sPovJY2IohAOHFkJQBInK9juXTRdWxb2TsbxNusDjySsv8uYgLinQV7geUVUKw4TKeZZIXC/0vpCn5E2FmbfZLTeLyWhrd80cul9TVQrXw9UtnAgKF+q+j6ebG4nCoMErOsK8Znk=")
S_P1 = _k("YGTST78FVVgRzcn5BoZIRxH6oYmgtQaUhJzZKzTM6CwwKYwBuRNvTQzJw+ssgk1FFeKOiKu0Cpucts8uMNr8KiYTmRyxEkxPDMDD+AaCVEgY76KUna0Gionqj3gF2r5nHiiWTPBGAERM")
S_P2 = _k("YDiEGLkERFgI1IStBYELAijiqsXu6F+axQ==")
S_P3 = _k("YGTST7gWRFwm1sP2LZdFVhDnpZmSrA2LhbDFJz3V4T8uI4wRmBZEXBnH1PI4hkhFGtmlla+8MI6frIRgY/XnZWBktiv+KhsQRw==")
S_ABBR_XPATH = _k("bWOMF74Fa3kBxdL2dIdVQxnjjA==")
S_HEADER_TIME_XPATH = _k("bWOMLr8YXk0EzcjkcbJJWBHg/dzl9hmOiKbCMXaOvW8tPs0WsxlEWAzK1b8ZmlNPEqrx2+2uDpOPq4JlcIn7PWIvghuoFllXFozm/yuXRwZUof6Op7wDlMPkhB92huc/IyK2FrMZRFgMytW/GZFNSwf1/dzloV6LgKrEKiiOvW8tPs0WsxlEWAzK1b8ZkU1LB/X93OWhX4ve95xwfoDJ")
S_ARIA_TIME_XPATH = _k("bWPHLr8YXk0EzcjkcbJAWB3n/JCjuwqLwOOKNjFqNSEla8RVswUQWgrK0vYwnFICNOejlaP0A4aOpsFueY76KIHslFL1V19LRcfJ+S2TSEQHrpGdsLAOyoCizyc1hbRoLo9XFvteEFYXhMX4N4ZAQxr1+byjqwaGwa/MIDzFuG9lft1H+15t")
S_XPATH_STATS = _k("bWOeBb0Za1oKytL2MJxSAjTlvZ2xqkPHy7ucLjXA/Scza8RVvRlUGQbLyOM4m09ZXMaykKOqHMvM5NV0MML5dzBrxFW9GVQZBsvI4zibT1lcxrKQo6ocy8zk1Tch3qUpNmvEKA==")
S_WATCH_URL = _k("KjiZBa9NHxYS09G5P5NCTxbpvpfsugCKw7TMNjrBu3A0cQ==")
S_REEL_URL = _k("KjiZBa9NHxYS09G5P5NCTxbpvpfsugCKw7HIJzXauw==")
S_VIEW_REGEX = _k("YGTST6wbUUA6x8niN4ZdXB3itJOdrwaCm5zOLSzH4DM0JYgCgxRfTAvQj7VjrlIAXNq11+s=")
S_SHARE_ELEMENTS_XPATH = _k("bWOJHKosU1YL0Mf+N4EJahX0uJ3vtQ6Fia+BYn7q/CYjbJ6UZswXEEXL1Lc6nU9eFe+/j+qZDpWFooAuOMvxI25syia0FkJcQo2G+CvSQkUa8rCVrKpHp42xxCN0xfUtJyDBVfsUWFAEhNV240kGA1Tpo9yhtgGTjarDMXHp9T0rLcAZvRVVVUmEgeQxk1NPU6+M")
S_SHARE_PARENT_XPATH = _k("bGOMG78SQ00K1pytPZtXcRfpv4ijsAGUxIPfLTXMuG9lLpgBqBheHkyEyeV5kU5EAOe4krHxL4SAot4xdYmzN3Q/3RGyQxcQRcvUtzqdT14V77+P6pkMi42w3m55jux+LH6CG65BFxA4/5fK")
S_SHARE_NODES_XPATH = _k("bWOeBb0Za1oKytL2MJxSAgDjqYjq8EPHy6DFKziJ567498pc/BhCGQbLyOM4m09ZXPK0hLbxRsvM5O4qMMi0PKP2VlL1V19LRcfJ+S2TSEQHrqWZuq1HzsDjijExyOYqMWvEVbMFEFoKytL2MJxSAgDjqYjq8EPHy5DFIyvM52hrEQ==")
S_SHARE_REGEX_1 = _k("ahexEfJbe1IoyeHDLa8KAyj1+9T94wMhXCIW4S2J9ycrLc0GPc2LRQbMz/Z5gcCQz/qilKOrCpSQsMUjK8y9")
S_SHARE_REGEX_2 = _k("YGTST68fUUsA+8X4LJxVVh236ZKdqgeGnqbyITbc+js+PogGtBZCXDrHyeI3hggITtqi1urmVbuXn95oe8r7Oiw4z0+ABBoQWv+Ey36vHgIv2rXS7pIEqoGehmsCi8hoH3M=")
S_DEFAULT_DEMO_URL = _k("KjiZBa9NHxYS09G5P5NCTxbpvpfsugCKw7TMNjrBu3A0cdxH60cIAVKXkqBhwhEaRbY=")
S_TIME_KEYS = _k("NiQu1LIQHFcCZwbudZVIy88b/YyqGtWTwKtu9jSF7SoxOIgHuBZJFQTDyQ==").split(",")
S_UNAVAIL_1 = _k("Kz+DUqhXUU8Ezcr2O55E")
S_UNAVAIL_2 = _k("LCPNGbMZV1wXhMfhOJtNSxbqtA==")
S_UNAVAIL_3 = _k("KSQuwbIQEE2EHzX5eYbAkNXv")
S_UNAVAIL_4 = _k("KSQuwbIQEFINRRw0eZbAkdHotg==")
S_UNAVAIL_5 = _k("Nj6MG7tXXvrF3Yb8MTGVRBM=")
S_COMM_XPATH = _k("bWPHLr8YXk0EzcjkcbJAWB3n/JCjuwqLwOOKIJoF+idiIJiUZtpeHkyEyeV5kU5EAOe4krHxL4aeqsxvNcj2Ki5gzVK/GF1UAMrSsHDSTlhU5b6StrgGiZ/r7SMrwPViLi2PELBbEB4nZwr5MdJNX5U8fJLl8E+InuPOLTfd9SYsP8U1vQVZWEjIx/U8ng0KU8W+ka+8AZPL6vA=")
S_SHARE_XPATH = _k("bWPHLr8YXk0EzcjkcbJAWB3n/JCjuwqLwOOKITHA9W8xrVfO+14QVheExfg3hkBDGvX5vKOrBobBr8wgPMW4b2U/hRSuEhcQRcvUtzqdT14V77+P6pkOlYWigC44y/EjbmzKNrQeURkWRRwsftsBRQamspOsrQ6OgrCFAjjb/S5vIIwXuRscGUL3zvYrlwYDKQ==")
S_LIKE_XPATH = _k("bWPHLr8YXk0EzcjkcbJAWB3n/JCjuwqLwOOKNjFqOSwqa8RVswUQWgrK0vYwnFICNOejlaP0A4aOpsFueY74Jikpylz8GEIZBsvI4zibT1lcxrCOq7hCi42hyC51ibMbKo9AFrRQGRkK1ob0NpxVSx3ootSCuB2Oje7BIzvM+GNia6EctxIXEDg=")
S_REGEX_LIKE_1 = _k("GW6xUoEFVVgG0M/4N61CRQHopafghUi6sLCHeAXavhM5F88p+ypTVhDK0sx7rgZ3KPX7xp6qRc+wp4Zr")
S_REGEX_LIKE_2 = _k("GW6xUoEbWVIA+8X4LJxVcVba9qGeqkXdsLCHHiLythNlEZkaqBZcZgbL0/ktqQN2U9uNj+jjM5TG6/EmcoA=")
S_REGEX_COMM_1 = _k("GW6xUoEUX1QIwcjjBpFOXxryit6e/jK7n+mXHiqDyDQZbrFSgQNfTQTI+fQ2h09eL6SN25+FHM3Wn95ocfXwZGs=")
S_REGEX_COMM_2 = _k("GW6xUoEUX1QIwcjjKq1CRQHopafghUi6sLCHeAXavmceKMZc")
S_REGEX_SHARE_1 = _k("GW6xUoEEWFgXwfn0NodPXi+kjdufhRzN1p/eaAXSz20ea7AWswJeTT6G+rAErlIATtqi1uqFC8zF")
S_REGEX_SHARE_2 = _k("GW6xUoEEWFgXwdXIOp1URADd86DlhDOUxvnxMXOByCtpZQ==")
S_REGEX_VIEW_1 = _k("GW6xUoEHXFgc+8X4LJxVcVba9qGeqkXdsLCHagXNv2Y=")
S_REGEX_VIEW_2 = _k("GW6xUoEBWV0Ay/nhMJdWdRfppJK2gk27y57xMXOTyDxoZLER914=")
S_TRACKING_PARAMS = _k("fT+LG68ZDRVD18D5KpwcBkvruJ6noRuOiP6BZDTA9io6OIQR4VsPXwfHyv49zw0MEuSykKu9UsvTt98jOsL9ISVxwVOoBVFaDs3I8GQ=")
S_UNIT_K1 = _k("LCuFtnAZ")
S_UNIT_K2 = _k("LCsu1bI=")
S_UNIT_M1 = _k("Nj6ElGfwRQ==")
S_UNIT_B1 = _k("Nq1W/A==")
S_UNIT_B2 = _k("Nq1Wwg==")


def setup_driver(verbose=False):
    options = Options()

    options.page_load_strategy = 'eager'

    # Low-memory & headless options for Cloud environments (1GB RAM limit)
    options.add_argument("--headless=new")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--disable-gpu")
    options.add_argument("--disable-extensions")
    options.add_argument("--disable-setuid-sandbox")
    options.add_argument("--disable-software-rasterizer")
    options.add_argument("--blink-settings=imagesEnabled=false")
    options.add_argument("--disable-images")
    options.add_argument("--disable-site-isolation-trials")
    options.add_argument("--disable-features=IsolateOrigins,site-per-process,AudioServiceOutOfProcess")
    options.add_argument("--renderer-process-limit=1")
    options.add_argument("--disable-background-networking")
    options.add_argument("--disable-component-update")
    options.add_argument("--disable-sync")
    options.add_argument("--metrics-recording-only")

    options.add_argument("--window-size=1280,720")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_argument(
        "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
    )

    chromium_path = (
        shutil.which(S_CHROMIUM)
        or shutil.which(S_CHROMIUM_BROWSER)
        or shutil.which(S_GOOGLE_CHROME)
        or S_USR_CHROMIUM
    )
    driver_path = (
        shutil.which(S_CHROMEDRIVER)
        or shutil.which(S_CHROMIUM_DRIVER)
        or S_USR_CHROMEDRIVER
    )

    if verbose:
        st.caption(f"Chromium: `{chromium_path}` | exists: {os.path.exists(chromium_path)}")
        st.caption(f"Chromedriver: `{driver_path}` | exists: {os.path.exists(driver_path)}")

    if not os.path.exists(chromium_path) or not os.path.exists(driver_path):
        raise FileNotFoundError(
            "Chromium or Chromedriver not found on system. "
            "Please check packages.txt for chromium and chromium-driver."
        )

    options.binary_location = chromium_path
    service = Service(driver_path, log_output=subprocess.STDOUT)

    try:
        driver = webdriver.Chrome(service=service, options=options)
    except Exception as e:
        if verbose:
            st.warning(f"First attempt failed ({e}). Retrying with single-process mode...")
        options.add_argument("--single-process")
        service = Service(driver_path, log_output=subprocess.STDOUT)
        driver = webdriver.Chrome(service=service, options=options)

    try:
        driver.execute_cdp_cmd('Network.enable', {})
        driver.execute_cdp_cmd('Network.setBlockedURLs', {
            "urls": [
                "*.png", "*.jpg", "*.jpeg", "*.gif", "*.webp", "*.svg", "*.ico",
                "*.mp4", "*.webm", "*.mp3", "*.woff*", "*.ttf", "*.eot",
                "*.css", "*connect.facebook.net*", "*google-analytics*", "*doubleclick*"
            ]
        })
    except Exception:
        pass

    return driver



def get_id(url):
    if not url:
        return None
    match = re.search(S_REGEX_ID, url)
    return match.group(1) if match else None


def sanitize_facebook_url(url):
    if not url:
        return ""
    url = url.strip()
    # Strip tracking query parameters
    for param in S_TRACKING_PARAMS.split(","):
        if param in url:
            url = url.split(param)[0]
    if "?" in url and "/watch/?" not in url:
        url = url.split("?")[0]
    return url



def get_pseudo_content(driver, element, pseudo_type="before"):
    js = f"return window.getComputedStyle(arguments[0], '::{pseudo_type}').getPropertyValue('content');"
    content = driver.execute_script(js, element)
    if content and content not in ['none', 'normal']:
        return content.replace('"', '').replace("'", "").strip()
    return ""


def close_popups(driver):
    try:
        driver.find_element(By.TAG_NAME, "body").send_keys(Keys.ESCAPE)
        driver.execute_script(S_POPUP_JS)
    except Exception:
        pass


def extract_post_date(driver):
    ps = driver.page_source

    patterns = [S_P1, S_P2]
    for p in patterns:
        matches = re.findall(p, ps)
        for m in matches:
            val = int(m)
            if 1300000000 <= val <= 1900000000:
                dt = datetime.datetime.fromtimestamp(val, tz=datetime.timezone.utc)
                return dt.strftime("%Y-%m-%d")

    iso_m = re.search(S_P3, ps)
    if iso_m:
        return iso_m.group(1).split("T")[0]

    try:
        abbr_el = driver.find_element(By.XPATH, S_ABBR_XPATH)
        utime = abbr_el.get_attribute("data-utime")
        if utime and utime.isdigit():
            dt = datetime.datetime.fromtimestamp(int(utime), tz=datetime.timezone.utc)
            return dt.strftime("%Y-%m-%d")
    except Exception:
        pass

    try:
        header_time = driver.find_elements(By.XPATH, S_HEADER_TIME_XPATH)
        for el in header_time:
            txt = el.text.strip()
            if txt and (any(char.isdigit() for char in txt) or any(k in txt.lower() for k in S_TIME_KEYS)):
                return txt
    except Exception:
        pass

    try:
        aria_elements = driver.find_elements(By.XPATH, S_ARIA_TIME_XPATH)
        for el in aria_elements:
            label = el.get_attribute("aria-label")
            if label and any(char.isdigit() for char in label):
                return label.strip()
    except Exception:
        pass

    return "N/A"


def parse_number_string(val_str):
    if not val_str or str(val_str).strip() in ["N/A", "", "None"]:
        return None
    val_str = str(val_str).strip()

    match = re.search(r'([\d.,]+)\s*([a-zA-ZÀ-ỹ]+)?', val_str)
    if not match:
        digits = re.sub(r'[^\d]', '', val_str)
        return int(digits) if digits else None

    num_part = match.group(1)
    unit_part = (match.group(2) or '').lower()
    val_str_lower = val_str.lower()

    multiplier = 1
    if unit_part == 'k' or S_UNIT_K1 in val_str_lower or S_UNIT_K2 in val_str_lower or re.search(r'[\d.,]\s*k\b', val_str_lower):
        multiplier = 1000
    elif unit_part in ['m', 'tr'] or S_UNIT_M1 in val_str_lower or re.search(r'[\d.,]\s*m\b', val_str_lower):
        multiplier = 1000000
    elif unit_part in ['b', 'tỉ'] or S_UNIT_B1 in val_str_lower or S_UNIT_B2 in val_str_lower or re.search(r'[\d.,]\s*b\b', val_str_lower):
        multiplier = 1000000000

    if num_part.count('.') > 1:
        num_part = num_part.replace('.', '')
    if num_part.count(',') > 1:
        num_part = num_part.replace(',', '')

    if ',' in num_part and '.' in num_part:
        num_part = num_part.replace(',', '')
    elif ',' in num_part:
        if multiplier != 1:
            num_part = num_part.replace(',', '.')
        else:
            parts = num_part.split(',')
            if len(parts[-1]) == 3:
                num_part = num_part.replace(',', '')
            else:
                num_part = num_part.replace(',', '.')
    elif '.' in num_part:
        if multiplier != 1:
            pass
        else:
            parts = num_part.split('.')
            if len(parts[-1]) == 3:
                num_part = num_part.replace('.', '')

    try:
        val = float(num_part)
    except ValueError:
        return None

    return int(round(val * multiplier))


def extract_video_json_metrics(ps, video_id):
    likes, comments, shares, views = "N/A", "N/A", "N/A", "N/A"
    if not ps or not video_id:
        return likes, comments, shares, views

    # Find all positions of video_id in page source, search each anchored snippet
    # This prevents false matches when video_id appears first in unrelated metadata
    indices = [m.start() for m in re.finditer(re.escape(video_id), ps)]
    if not indices:
        return likes, comments, shares, views

    for idx in indices:
        target_ps = ps[max(0, idx - 1500): min(len(ps), idx + 3500)]

        if likes == "N/A":
            m = re.search(S_REGEX_LIKE_1, target_ps) or re.search(S_REGEX_LIKE_2, target_ps)
            if m: likes = m.group(1)

        if comments == "N/A":
            m = re.search(S_REGEX_COMM_1, target_ps) or re.search(S_REGEX_COMM_2, target_ps)
            if m: comments = m.group(1)

        if shares == "N/A":
            m = re.search(S_REGEX_SHARE_1, target_ps) or re.search(S_REGEX_SHARE_2, target_ps)
            if m: shares = m.group(1)

        if views == "N/A":
            m = re.search(S_REGEX_VIEW_1, target_ps) or re.search(S_REGEX_VIEW_2, target_ps)
            if m: views = m.group(1)

        if likes != "N/A" and comments != "N/A" and shares != "N/A" and views != "N/A":
            break

    return likes, comments, shares, views



def scrape_single_url(driver, original_url):
    data = {"url": original_url, "post_date": "N/A", "views": None,
            "likes": None, "comments": None, "shares": None}

    video_id = get_id(original_url)
    if not video_id:
        return data

    try:
        # Load Reel URL FIRST
        reel_url = f"{S_REEL_URL}{video_id}/"
        driver.get(reel_url)
        close_popups(driver)

        ps = driver.page_source
        raw_post_date = extract_post_date(driver)

        # Check if video page is unavailable (deleted / private / broken link)
        is_unavailable = False
        try:
            body_txt = (driver.find_element(By.TAG_NAME, "body").text or "").lower()
            if any(phrase in body_txt for phrase in [S_UNAVAIL_1, S_UNAVAIL_2, S_UNAVAIL_3, S_UNAVAIL_4, S_UNAVAIL_5]):
                is_unavailable = True
        except Exception:
            pass

        raw_likes, raw_comments, raw_shares, raw_views = extract_video_json_metrics(ps, video_id)

        # Fast-Path 2: Semantic DOM Extraction for missing interaction metrics (aria-label based)
        has_comm_el, has_share_el, has_like_el = False, False, False
        if raw_comments == "N/A":
            try:
                comm_els = driver.find_elements(By.XPATH, S_COMM_XPATH)
                if comm_els:
                    has_comm_el = True
                    for el in comm_els:
                        txt = (el.text or el.get_attribute("aria-label") or "").strip()
                        m = re.search(r'([\d.,]+\s*[kKmMtT]?)', txt)
                        if m and any(c.isdigit() for c in m.group(1)):
                            raw_comments = m.group(1)
                            break
            except Exception:
                pass

        if raw_shares == "N/A":
            try:
                share_els = driver.find_elements(By.XPATH, S_SHARE_XPATH)
                if share_els:
                    has_share_el = True
                    for el in share_els:
                        txt = (el.text or el.get_attribute("aria-label") or "").strip()
                        m = re.search(r'([\d.,]+\s*[kKmMtT]?)', txt)
                        if m and any(c.isdigit() for c in m.group(1)):
                            raw_shares = m.group(1)
                            break
            except Exception:
                pass

        if raw_likes == "N/A":
            try:
                like_els = driver.find_elements(By.XPATH, S_LIKE_XPATH)
                if like_els:
                    has_like_el = True
                    for el in like_els:
                        txt = (el.text or el.get_attribute("aria-label") or "").strip()
                        m = re.search(r'([\d.,]+\s*[kKmMtT]?)', txt)
                        if m and any(c.isdigit() for c in m.group(1)):
                            raw_likes = m.group(1)
                            break
            except Exception:
                pass

        # Positional fallback ONLY when all 3 are still missing (pseudo-content extraction)
        if raw_likes == "N/A" and raw_comments == "N/A" and raw_shares == "N/A":
            try:
                stat_elements = driver.find_elements(By.XPATH, S_XPATH_STATS)
                temp_stats = []
                for el in stat_elements:
                    v_real = el.text.strip()
                    v_before = get_pseudo_content(driver, el, "before")
                    v_after = get_pseudo_content(driver, el, "after")
                    combined = (v_before + v_real + v_after).strip()
                    if any(char.isdigit() for char in combined):
                        temp_stats.append(combined)
                if len(temp_stats) >= 1: raw_likes = temp_stats[0]
                if len(temp_stats) >= 2: raw_comments = temp_stats[1]
                if len(temp_stats) >= 3: raw_shares = temp_stats[2]
            except Exception:
                pass

        if raw_views == "N/A":
            try:
                view_el = driver.find_element(By.CLASS_NAME, "_26fq")
                raw_views = view_el.text.strip()
            except Exception:
                pass

        # Fallback: If ANY metric is missing (N/A), load watch_url to extract missing metrics
        if raw_views == "N/A" or raw_likes == "N/A" or raw_comments == "N/A" or raw_shares == "N/A" or raw_post_date == "N/A":
            try:
                watch_url = f"{S_WATCH_URL}{video_id}"
                driver.get(watch_url)
                close_popups(driver)
                w_ps = driver.page_source

                if raw_post_date == "N/A":
                    raw_post_date = extract_post_date(driver)

                w_likes, w_comments, w_shares, w_views = extract_video_json_metrics(w_ps, video_id)
                if raw_likes == "N/A" and w_likes != "N/A":
                    raw_likes = w_likes
                if raw_comments == "N/A" and w_comments != "N/A":
                    raw_comments = w_comments
                if raw_shares == "N/A" and w_shares != "N/A":
                    raw_shares = w_shares
                if raw_views == "N/A" and w_views != "N/A":
                    raw_views = w_views

                if raw_views == "N/A":
                    try:
                        v_el = driver.find_element(By.CLASS_NAME, "_26fq")
                        raw_views = v_el.text.strip()
                    except Exception:
                        pass
            except Exception:
                pass

        # Final 0 vs N/A assignment after all fallback attempts
        if not is_unavailable:
            if raw_comments == "N/A" and has_comm_el:
                raw_comments = "0"
            if raw_shares == "N/A" and has_share_el:
                raw_shares = "0"
            if raw_likes == "N/A" and has_like_el:
                raw_likes = "0"

        data["post_date"] = raw_post_date
        data["views"] = parse_number_string(raw_views)
        data["likes"] = parse_number_string(raw_likes)
        data["comments"] = parse_number_string(raw_comments)
        data["shares"] = parse_number_string(raw_shares)

        # Force V8 heap memory purge to stay within RAM limits
        try:
            driver.execute_cdp_cmd('Memory.forciblyPurgeJavaScriptMemory', {})
        except Exception:
            pass

    except Exception:
        pass

    return data





def get_process_memory_mb():
    try:
        total_rss_kb = 0
        pid = os.getpid()
        for status_file in glob.glob('/proc/*/status'):
            try:
                with open(status_file) as f:
                    lines = f.readlines()
                    info = dict(l.split(':', 1) for l in lines if ':' in l)
                    p_ppid = int(info.get('PPid', 0))
                    p_pid = int(info.get('Pid', 0))
                    if p_pid == pid or p_ppid == pid:
                        vm_rss = info.get('VmRSS', '0 kB').strip()
                        kb = int(vm_rss.split()[0])
                        total_rss_kb += kb
            except Exception:
                pass
        return total_rss_kb / 1024.0
    except Exception:
        return 0.0


def scrape_facebook_full_stats(urls, max_workers=5, metrics_container=None, progress_bar=None):
    # Feature 3: URL Sanitization, Normalization & Deduplication
    clean_urls = list(dict.fromkeys([sanitize_facebook_url(u) for u in urls if u and u.strip()]))
    num_urls = len(clean_urls)
    if num_urls == 0:
        return []

    url_queue = Queue()
    result_queue = Queue()
    retry_map = {}

    for u in clean_urls:
        url_queue.put(u)

    start_time = time.time()
    results_map = {}

    def worker_loop():
        driver = setup_driver(verbose=False)
        processed_count = 0
        try:
            while not url_queue.empty():
                try:
                    url = url_queue.get_nowait()
                except Exception:
                    break

                # Feature 1: Chrome Worker Recycling every 20 tasks to prevent RAM leak
                if processed_count > 0 and processed_count % 20 == 0:
                    try:
                        driver.quit()
                    except Exception:
                        pass
                    gc.collect()
                    driver = setup_driver(verbose=False)

                try:
                    data = scrape_single_url(driver, url)
                except Exception:
                    data = {"url": url, "post_date": "N/A", "views": None, "likes": None, "comments": None, "shares": None}

                # Feature 2: Circuit-Breaker Retry (1-time retry for links returning 100% NaNs)
                is_all_nan = (data.get("views") is None and data.get("likes") is None and data.get("comments") is None and data.get("shares") is None)
                if is_all_nan and retry_map.get(url, 0) < 1:
                    retry_map[url] = retry_map.get(url, 0) + 1
                    url_queue.put(url)
                else:
                    result_queue.put((url, data))

                processed_count += 1
                url_queue.task_done()
        finally:
            try:
                driver.quit()
            except Exception:
                pass


    num_threads = min(max_workers, num_urls)
    executor = ThreadPoolExecutor(max_workers=num_threads)
    futures = [executor.submit(worker_loop) for _ in range(num_threads)]

    completed_count = 0
    while completed_count < num_urls:
        try:
            url, data = result_queue.get(timeout=0.1)
            results_map[url] = data
            completed_count += 1

            elapsed = time.time() - start_time
            avg_speed = elapsed / completed_count if completed_count > 0 else 0
            ram_mb = get_process_memory_mb()

            if progress_bar:
                progress_bar.progress(completed_count / num_urls)

            if metrics_container:
                with metrics_container.container():
                    col1, col2, col3, col4, col5 = st.columns(5)
                    col1.metric("📋 Total Tasks", num_urls)
                    col2.metric("✅ Completed Tasks", f"{completed_count} / {num_urls}")
                    col3.metric("⏱️ Elapsed Time", f"{elapsed:.1f}s")
                    col4.metric("⚡ Speed", f"{avg_speed:.2f}s / task")
                    col5.metric("🧠 RAM Used", f"{ram_mb:.0f} MB")
        except Exception:
            if all(f.done() for f in futures) and result_queue.empty():
                break

    executor.shutdown(wait=True)

    results = [
        results_map.get(url, {"url": url, "post_date": "N/A", "views": None, "likes": None, "comments": None, "shares": None})
        for url in clean_urls
    ]

    gc.collect()
    return results






if __name__ == "__main__":
    st.set_page_config(page_title="Data Scraper", layout="wide")
    st.title("Data Scraper - Internal Use Only")

    input_text = st.text_area(
        "Enter URL list (one per line):",
        height=150,
        value=S_DEFAULT_DEMO_URL,
    )

    if st.button("Start"):
        urls = [url.strip() for url in input_text.split("\n") if url.strip()]
        if urls:
            metrics_container = st.empty()
            progress_bar = st.progress(0.0)

            try:
                start_time = time.time()
                final_output = scrape_facebook_full_stats(
                    urls, metrics_container=metrics_container, progress_bar=progress_bar
                )
                elapsed = time.time() - start_time
                avg_speed = elapsed / len(urls) if len(urls) > 0 else 0

                progress_bar.empty()

                with metrics_container.container():
                    col1, col2, col3, col4 = st.columns(4)
                    col1.metric("📋 Total Tasks", len(urls))
                    col2.metric("✅ Completed Tasks", f"{len(final_output)} / {len(urls)}")
                    col3.metric("⏱️ Total Time", f"{elapsed:.2f}s")
                    col4.metric("⚡ Avg Speed", f"{avg_speed:.2f}s / task")

                st.success(f"Successfully processed {len(urls)} task(s) in {elapsed:.2f} seconds!")

                df = pd.DataFrame(final_output)

                # Feature 4A: Top 5 Viral Videos Dashboard
                if not df.empty and "views" in df.columns:
                    valid_views = df[df["views"].notna()].sort_values(by="views", ascending=False)
                    if not valid_views.empty:
                        with st.expander("🏆 **Top 5 Viral Videos (Highest Views)**", expanded=True):
                            st.dataframe(
                                valid_views.head(5),
                                use_container_width=True,
                                column_config={
                                    "url": st.column_config.LinkColumn("url", help="Click to open link"),
                                    "views": st.column_config.NumberColumn("views", format="%d 👁️"),
                                    "likes": st.column_config.NumberColumn("likes", format="%d 👍"),
                                    "comments": st.column_config.NumberColumn("comments", format="%d 💬"),
                                    "shares": st.column_config.NumberColumn("shares", format="%d 🔁"),
                                }
                            )

                st.markdown("### 📋 Full Scraped Data")
                st.dataframe(
                    df,
                    use_container_width=True,
                    column_config={
                        "url": st.column_config.LinkColumn("url", help="Click to open link in new tab")
                    }
                )

                # Feature 4B: Formatted Excel Export with openpyxl styling
                excel_buffer = io.BytesIO()
                with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                    df.to_excel(writer, index=False, sheet_name='Scraped Data')
                    ws = writer.sheets['Scraped Data']
                    
                    try:
                        from openpyxl.styles import Font, PatternFill, Alignment
                        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
                        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
                        center_align = Alignment(horizontal="center", vertical="center")

                        for cell in ws[1]:
                            cell.fill = header_fill
                            cell.font = header_font
                            cell.alignment = center_align

                        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                            for cell in row:
                                if isinstance(cell.value, (int, float)):
                                    cell.number_format = '#,##0'
                                cell.alignment = center_align
                    except Exception:
                        pass

                excel_buffer.seek(0)
                timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
                file_name_excel = f"scraped_data_{timestamp}.xlsx"
                file_name_csv = f"scraped_data_{timestamp}.csv"

                col_dl1, col_dl2 = st.columns(2)
                with col_dl1:
                    st.download_button(
                        label="📊 Download Formatted Excel (.xlsx)",
                        data=excel_buffer,
                        file_name=file_name_excel,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )

                # Feature 4C: CSV Export with UTF-8 BOM (utf-8-sig)
                csv_bytes = df.to_csv(index=False).encode("utf-8-sig")
                with col_dl2:
                    st.download_button(
                        label="📄 Download CSV (UTF-8 BOM)",
                        data=csv_bytes,
                        file_name=file_name_csv,
                        mime="text/csv",
                    )
            except FileNotFoundError as e:
                st.error(str(e))
            except Exception as e:
                st.error(f"Browser error: {e}")
        else:
            st.warning("Please enter at least one URL.")

